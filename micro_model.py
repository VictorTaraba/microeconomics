from openpyxl import load_workbook

workbook = load_workbook(filename="micro_d.xlsx")
sheet = workbook.active
list_first = []
for i in range(1,37):
	list_first.append(sheet.cell(row=1,column=i).value)
workbook.close()

N = len(list_first)
count_a = float(round(list_first.count('a')/N,2))
count_b = float(round(list_first.count('b')/N,2))
count_c = float(round(list_first.count('c')/N,2))

def absolute_majority(list_first):
	text=''
	if count_a>0.5:
		print('yes')
		text="Переможець: a."
	elif count_b>0.5:
		text="Переможець: b."
	elif count_c>0.5:
		text="Переможець: c."
	else:
		text="Не вдалося визначити переможця"

	f = open('results.txt', 'w')
	f.write("--- Абсолютна більшість ---"+'\n')
	f.write("Частка голосів за a: "+str(count_a)+\
		'.'+'\n'+"Частка голосів за b: "+str(count_b)+\
		'.'+'\n'+"Частка голосів за c: "+str(count_c)+\
		'.'+'\n'+text+'\n'*2)
	f.close()

def relative_majority(list_first):
	text=''
	list_t = [count_a, count_b, count_c]
	max_i = list_t.index(max(list_t))
	if max_i == 0:
		text="Переможець: a."
	elif max_i == 1:
		text="Переможець: b."
	elif max_i == 2:
		text="Переможець: c."

	f = open('results.txt', 'a+')
	f.write("--- Відносна більшість ---"+'\n')
	f.write("Частка голосів за a: "+str(count_a)+\
		'.'+'\n'+"Частка голосів за b: "+str(count_b)+\
		'.'+'\n'+"Частка голосів за c: "+str(count_c)+\
		'.'+'\n'+text+'\n'*2)
	f.close()

def Borda_count(list_first):
	a_result, b_result, c_result = 0, 0, 0
	for i in range(1,37):
		list_r = []
		for j in range(1,4):
			list_r.append(sheet.cell(row=j,column=i).value)

		if list_r.index("a") == 0:
			add = 2
		elif list_r.index("a") == 1:
			add = 1
		else:
			add=0
		a_result=a_result+add

		if list_r.index("b") == 0:
			add = 2
		elif list_r.index("b") == 1:
			add = 1
		else:
			add=0
		b_result=b_result+add

		if list_r.index("c") == 0:
			add = 2
		elif list_r.index("c") == 1:
			add = 1
		else:
			add=0
		c_result=c_result+add

	list_m = [a_result, b_result, c_result]
	max_i = list_m.index(max(list_m))
	if (max_i == 0) and (list_m[0]!=list_m[1]) and (list_m[0]!=list_m[2]):
		text="Переможець: a."
	elif (max_i == 1) and (list_m[1]!=list_m[0]) and (list_m[1]!=list_m[2]):
		text="Переможець: b."
	elif (max_i == 2) and (list_m[2]!=list_m[1]) and (list_m[2]!=list_m[0]):
		text="Переможець: c."
	else:
		text="Не вдалося визначити переможця."

	f = open('results.txt', 'a+')
	f.write("--- Правило Борда ---"+'\n')
	f.write("Балів у a: "+str(a_result)+\
		'.'+'\n'+"Балів у b: "+str(b_result)+\
		'.'+'\n'+"Балів у c: "+str(c_result)+\
		'.'+'\n'+text+'\n'*2)
	f.close()

absolute_majority(list_first)
relative_majority(list_first)
Borda_count(list_first)

